import time

from uart import *

from colorama import Fore, Style

from loadcode import *

import os
from openpyxl import load_workbook
from openpyxl import Workbook

from datetime import datetime

from openpyxl.styles import PatternFill
from openpyxl.styles import Font

import winsound
import keyboard

def Buzzer_pip():
    winsound.Beep(1000, 200)
    
def Buzzer_error():
    winsound.Beep(440, 1000)
    time.sleep(0.1)
    winsound.Beep(440, 1000)
    time.sleep(0.1)
    winsound.Beep(440, 1000)

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
bold_font = Font(bold=True)  # Định dạng chữ đậm


TEST_OK         =   0x59
TEST_ERROR  	=	0x4E
TEST_ERROR_1	=	0x5E
TEST_ERROR_2	=	0x6E

Array_print = "Mang in ra"

def In_loi(data):
    print(Fore.RED + data + Style.RESET_ALL)
    
def In_ketqua(data):
    print(Fore.GREEN + data + Style.RESET_ALL)
    
def Cal_Checksum(data):
    CS_b = 0
    for i in range(0,11):
        CS_b += data[i]
    CS_b = CS_b % 256
    return CS_b

def Led_process(_uart_var, firm_path_goc, log_path_goc, infor_mach):
    # Lưu log dữ liệu test của mạch ------------------------------------------ 
    folder_path = log_path_goc+ "\Led"

    part_infor = infor_mach.split('_')
    Header = part_infor[0]
    MachCanTest = part_infor[1]
    LoSanXuat = part_infor[2]
    Maso = part_infor[3]
    file_name = Maso + ".xlsx"
    
    solan = 1
    while True:
        file_path = os.path.join(folder_path, file_name)
        if os.path.exists(file_path):
            solan = solan + 1
            file_name = str(Maso) + "_Lan_" + str(solan) + ".xlsx"      
        else:
            # Nếu file không tồn tại, tạo file mới
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Result_Test"
            wb.save(file_path)
            # print(f"->File '{file_name}' đã được tạo.")
            break
    C12 = "-"
    C13 = "-"
    C14 = "-"
    C15 = "-"
    C16 = "-"
    C17 = "-"
    C18 = "-"
    C19 = "-"
    

    # Gửi lệnh reset mạch JIG-------------------------------------------------
    buffer_cmd = [0xAB, 0xCD, 0x00, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x21]
    _uart_var.send_data(buffer_cmd)
    time.sleep(0.5)
    
    vErrorCode = ""
    buffer_cmd = [0xAB, 0xCD, 0x10, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x31]
    # Send_CMD_to_JIG(port, buffer_cmd)
    _uart_var.send_data(buffer_cmd)

    while True:  
        vKetthuc = False
        data_res = _uart_var.read_data()
        if data_res:  
            CS_byte = Cal_Checksum(data_res)              
            if (data_res[0] == 0xEA) & (data_res[11] == CS_byte):
                # ID
                if data_res[1] == 0x10:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> ID mạch cần test đúng !\n"
                        In_ketqua(Array_print)
                        print("> Đang test trở kháng...")
                    else:
                        Array_print = "-> ERROR: Sai ID !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_ID_Led"
                        vKetthuc = True
                # Impedance
                elif data_res[1] == 0x11:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test trở kháng thành công!\n"
                        In_ketqua(Array_print)
                        print("> Đang test điện áp...")
                        C12 = "PASS"
                        C13 = "PASS"
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch LED lộ 5vBus !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_TroKhang_5vBus_Led"
                        vKetthuc = True
                        C12 = "ERROR"
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch LED lộ 3v3 !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_TroKhang_3v3_Led"
                        vKetthuc = True
                        C12 = "PASS"
                        C13 = "ERROR"
                # Voltage
                elif data_res[1] == 0x12:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test điện áp thành công!\n"
                        In_ketqua(Array_print)
                        C14 = "PASS"
                        C15 = "PASS"
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch LED lộ 5v0 !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_DienAp_5v0_Led"
                        vKetthuc = True
                        C14 = "ERROR"
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch LED lộ 3v3 !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_DienAp_3v0_Led"
                        vKetthuc = True
                        C14 = "PASS"
                        C15 = "ERROR"
                # St-Link
                elif data_res[1] == 0x13:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Đã kết nối St-Link, bắt đầu nạp code test !\n"
                        In_ketqua(Array_print)
                        # Nap code test
                        firm_path_led_test = firm_path_goc + "\led_board\led_test.hex"
                        kq_fw = flash_firmware(firm_path_led_test)
                        if kq_fw == 1:
                            Array_print = "-> Nạp firmware test thành công!\n"
                            In_ketqua(Array_print)
                            # Gửi lệnh kiểm tra code test
                            time.sleep(0.5)
                            buffer_cmd = [0xAB, 0xCD, 0x14, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x35]
                            _uart_var.send_data(buffer_cmd)
                            print("> Đang kiểm tra code test...")
                        else:
                            Array_print = "-> ERROR: Nạp firmware test thất bại !!!\n"
                            In_loi(Array_print)
                            C16 = "ERROR"
                            vErrorCode = "ERROR_NapCode_Led"
                            vKetthuc = True
                    elif data_res[2] == TEST_ERROR:
                        Array_print = "-> ERROR: Lỗi kết nối St-Link !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_STLink_Led"
                        vKetthuc = True
                # Code test
                elif data_res[1] == 0x14:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Kiểm tra nạp code test thành công!\n"
                        In_ketqua(Array_print)
                        C16 = "PASS"
                        print("> Đang test functions...")
                    elif data_res[2] == TEST_ERROR:
                        Array_print = "-> ERROR: Kiểm tra nạp code test thất bại !!!\n"
                        In_loi(Array_print)
                        C16 = "ERROR"
                        print(Fore.MAGENTA +"             Bạn có muốn nạp lại không?      \n"+ Style.RESET_ALL)
                        print("          1. Có                    2. Không  \n")
                        kq = int(input("> Nhập:"))
                        if kq == 1:
                            # Nap code test
                            firm_path_led_test = firm_path_goc + "\led_board\led_test.hex"
                            kq_fw = flash_firmware(firm_path_led_test)
                            if kq_fw == 1:
                                Array_print = "-> Nạp firmware test thành công!\n"
                                In_ketqua(Array_print)
                                # Gửi lại lệnh kiểm tra code test
                                time.sleep(0.5)
                                buffer_cmd = [0xAB, 0xCD, 0x14, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x35]
                                _uart_var.send_data(buffer_cmd)
                                print("> Đang kiểm tra code test")
                            else:
                                Array_print = "-> ERROR: Nạp firmware test thất bại !!!\n"
                                In_loi(Array_print)
                                C16 = "ERROR"
                                # Gửi lệnh reset Jig
                                time.sleep(0.5)
                                buffer_cmd = [0xAB, 0xCD, 0x00, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x21]
                                _uart_var.send_data(buffer_cmd)
                                vErrorCode = "ERROR_NapCode_Led"
                                vKetthuc = True
                        else:
                            # Gửi lệnh reset Jig
                            time.sleep(0.5)
                            buffer_cmd = [0xAB, 0xCD, 0x00, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x21]
                            _uart_var.send_data(buffer_cmd)
                            vErrorCode = "ERROR_NapCodeTest_Led"
                            vKetthuc = True
                # Function
                elif data_res[1] == 0x15:
                    if data_res[2] == TEST_OK:
                        C17 = "PASS"
                        #Full led
                        print(Fore.MAGENTA +"         Tất cả led có sáng không (23 led)?           \n"+ Style.RESET_ALL)
                        print("      1.Có                           2. Không\n")
                        Buzzer_pip()
                        led = int(input("> Nhập vào kết quả quan sát: "))
                        if led == 1:
                            Array_print = "-> Test Funtion thành công, bắt đầu nạp code chính !\n"
                            In_ketqua(Array_print)
                            C18 = "PASS"
                            # Nạp code chính
                            firm_path_led_test = firm_path_goc + "\led_board\led_main.hex"
                            kq_fw = flash_firmware(firm_path_led_test)
                            if kq_fw == 1:
                                Array_print = "-> Nạp firmware chính thành công!\n"
                                In_ketqua(Array_print)
                                # Gửi lệnh kiểm tra code chính
                                time.sleep(0.5)
                                buffer_cmd = [0xAB, 0xCD, 0x16, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x37]
                                _uart_var.send_data(buffer_cmd)
                                print("> Đang kiểm tra code chính")
                            else:
                                Array_print = "-> ERROR: Nạp firmware chính thất bại !!!\n"
                                In_loi(Array_print)
                                C19 = "ERROR"
                                vErrorCode = "ERROR_NapCode_Led"
                                vKetthuc = True
                        elif led == 2:
                            Array_print = "-> ERROR: Lỗi LED trên mạch !!!\n"
                            In_loi(Array_print)
                            C18 = "ERROR"
                            # Gửi lệnh reset Jig
                            time.sleep(0.5)
                            buffer_cmd = [0xAB, 0xCD, 0x00, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x21]
                            _uart_var.send_data(buffer_cmd)
                            vErrorCode = "ERROR_FullLed_Led"
                            vKetthuc = True
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi Mạch LED RS485 !!!\n"
                        In_loi(Array_print)
                        C17 = "ERROR"
                        vErrorCode = "ERROR_RS485_Led"
                        vKetthuc = True
                # Code chính
                elif data_res[1] == 0x16:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Kiểm tra nạp code chính thành công, ***MẠCH ĐẠT YÊU CẦU!\n"
                        In_ketqua(Array_print)
                        C19 = "PASS"
                        vErrorCode = "TEST_LED_OK"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR:
                        Array_print = "-> ERROR: Kiểm tra nạp code chính thất bại !!!\n"
                        In_loi(Array_print)
                        C19 = "ERROR"
                        print(Fore.MAGENTA +"             Bạn có muốn nạp lại không?      \n"+ Style.RESET_ALL)
                        print("          1. Có                    2. Không")
                        kq = int(input("> Nhập:"))
                        if kq == 1:
                            # Nạp code chính
                            firm_path_led_test = firm_path_goc + "\led_board\led_main.hex"
                            kq_fw = flash_firmware(firm_path_led_test)
                            if kq_fw == 1:
                                Array_print = "-> Nạp firmware chính thành công!\n"
                                In_ketqua(Array_print)
                                # Gửi lại lệnh kiểm tra code chính
                                time.sleep(0.5)
                                buffer_cmd = [0xAB, 0xCD, 0x16, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x37]
                                _uart_var.send_data(buffer_cmd)
                                print("> Đang kiểm tra code chính...")
                            else:
                                Array_print = "-> ERROR: Nạp firmware chính thất bại !!!\n"
                                In_loi(Array_print)
                                C19 = "ERROR"
                                vErrorCode = "ERROR_NapCode_Led"
                                vKetthuc = True
                                # Gửi lệnh reset Jig
                                time.sleep(0.1)
                                buffer_cmd = [0xAB, 0xCD, 0x00, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x21]
                                _uart_var.send_data(buffer_cmd)
                        else:
                            # Gửi lệnh reset Jig
                            time.sleep(0.1)
                            buffer_cmd = [0xAB, 0xCD, 0x00, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x21]
                            _uart_var.send_data(buffer_cmd)
                            vErrorCode = "ERROR_NapCodeChinh_Led"
                            vKetthuc = True
                else:
                    print("Nothing\n")
            else:
                print("-> Sai Cấu trúc bản tin Response !!!\n")
        
        if vKetthuc == True:
            break
        
        if keyboard.is_pressed('q'):
            vErrorCode = "LỖI QUY TRÌNH TEST"
            print("\n")
            print(Fore.RED + "> Đã nhấn Exit. Kết thúc test mạch!\n" + Style.RESET_ALL)
            break     
        
        
        
    # Lưu dữ liệu kết quả test--------------------------------------
    wb = load_workbook(file_path)
    ws = wb.active
    
    ws["A1"] = "Board's name"
    
    if vErrorCode != "ERROR_ID_Led":
        ws["B1"] = "LED - Led"
    else:
        ws["B1"] = "Sai loại mạch"
        ws["B1"].fill = red_fill

    ws["A2"] = "Design Company"
    ws["B2"] = "Viettel Post"
    ws["A3"] = "Factory's name"
    ws["B3"] = "Meiko"
    ws["A4"] = "Production Batch"
    ws["B4"] = LoSanXuat
    ws["A5"] = "Serial Number"
    ws["B5"] = Maso
    ws["A6"] = "Date"
    ws["A7"] = "Time"
        
    now = datetime.now()
    # Định dạng ngày và giờ
    ws["B6"] = now.strftime("%Y-%m-%d")  # Ngày hiện tại (YYYY-MM-DD)
    ws["B7"] = now.strftime("%H:%M:%S")  # Giờ hiện tại (HH:MM:SS)    
       
    ws["A8"] = "Number of times" 
    ws["B8"] =  solan
      
    ws["A9"] = "Result"   
    if vErrorCode == "TEST_LED_OK":
       ws["B9"] =  "PASS"
       ws["B9"].fill =  green_fill
    else:
       ws["B9"] =  "ERROR" 
       ws["B9"].fill =  red_fill
    ws["B9"].font =  bold_font
    
    ws["A11"] =  "No" 
    ws["B11"] =  "Test Function" 
    ws["C11"] =  "Result"
    
    ws["A12"] =  '1'
    ws["A13"] =  '2'
    ws["A14"] =  '3'
    ws["A15"] =  '4'
    ws["A16"] =  '5'
    ws["A17"] =  '6'
    ws["A18"] =  '7'
    ws["A19"] =  '8'
    
    ws["B12"] =  "Impedance path 5vBus"
    ws["B13"] =  "Impedance path 3v3"
    ws["B14"] =  "Voltage path 5v0"
    ws["B15"] =  "Voltage path 3v3"
    ws["B16"] =  "Load test program"
    ws["B17"] =  "RS485"
    ws["B18"] =  'On all led on board'
    ws["B19"] =  'Load main program'
    
    ws["C12"] =  C12
    ws["C13"] =  C13
    ws["C14"] =  C14
    ws["C15"] =  C15
    ws["C16"] =  C16
    ws["C17"] =  C17
    ws["C18"] =  C18
    ws["C19"] =  C19
    
    # Lưu file Excel
    wb.save(file_path)
    
    # Gửi lệnh reset Jig
    time.sleep(0.5)
    buffer_cmd = [0xAB, 0xCD, 0x00, 0x03, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x21]
    _uart_var.send_data(buffer_cmd)
    
    return vErrorCode