import time
import keyboard

from uart import *

from colorama import Fore, Style

from loadcode import *

import os
from openpyxl import load_workbook
from openpyxl import Workbook

from datetime import datetime

from openpyxl.styles import PatternFill
from openpyxl.styles import Font

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
bold_font = Font(bold=True)  # Định dạng chữ đậm


TEST_OK         =   0x59
TEST_ERROR  	=	0x4E
TEST_ERROR_1	=	0x5E
TEST_ERROR_2	=	0x6E
TEST_ERROR_3	=	0x7E
TEST_ERROR_4	=	0x8E
TEST_ERROR_5	=	0x9E
TEST_ERROR_6	=	0xAE

ERROR_RS485		=	0xE0
	
ERROR_OS_ADC_1_1		=	0xB0
ERROR_OS_ADC_2_1		=	0xB1
ERROR_OS_ADC_3_1		=	0xB2
ERROR_OS_ADC_4_1		=	0xB3

ERROR_OS_ADC_1_2		=	0xB4
ERROR_OS_ADC_2_2		=	0xB5
ERROR_OS_ADC_3_2		=	0xB6
ERROR_OS_ADC_4_2		=	0xB7
ERROR_OS_ADC_1_3		=	0xB8
ERROR_OS_ADC_2_3		=	0xB9
ERROR_OS_ADC_3_3		=	0xBA
ERROR_OS_ADC_4_3		=	0xBB

ERROR_MOSFET_1		=	0xE1
ERROR_MOSFET_2		=	0xE2
ERROR_MOSFET_3		=	0xE3
ERROR_MOSFET_4		=	0xE4

ERROR_NTC			=			0xE5
ERROR_FAN			=			0xE6


CMD_CHECK_ID_CT		    =	0x10
CMD_CHECK_IMPEDANCE	    =   0x20

CMD_CHECK_VOLTAGE_BAT	=	0x30
CMD_CHECK_VOLTAGE_CH	=	0x31

CMD_CHECK_ST_LINK	    =	0x40

CMD_CHECK_CODE_TEST	    =   0x50
CMD_CHECK_FUNCTIONS	    =   0x60

CMD_CHECK_CODE_MAIN     =	0x70

Array_print = "Mang in ra"

cmd_reset =     [0xAB, 0xCD , 0x00 , 0x06 , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0x24]

cmd_phase1_10 = [0xAB, 0xCD , 0x10 , 0x06 , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0x34]

cmd_phase2_50 = [0xAB, 0xCD , 0x50 , 0x06 , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0x74]

cmd_phase3_70 = [0xAB, 0xCD , 0x70 , 0x06 , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0xAA , 0x94]

def In_loi(data):
    print(Fore.RED + data + Style.RESET_ALL)
    
def In_ketqua(data):
    print(Fore.GREEN + data + Style.RESET_ALL)
    
def Cal_Checksum(data):
    CS_b = 0
    for i in range(0,11):
        CS_b += data[i]
        CS_b = CS_b % 256
    return CS_b

def Power_process(_uart_var, firm_path_goc, log_path_goc, infor_mach):
    # Lưu log dữ liệu test của mạch ------------------------------------------ 
    folder_path = log_path_goc+ "\Power"

    part_infor = infor_mach.split('_')
    Header = part_infor[0]
    MachCanTest = part_infor[1]
    LoSanXuat = part_infor[2]
    Maso = part_infor[3]
    file_name = Maso + ".xlsx"
    
    solan = 1
    while True:
        file_path = os.path.join(folder_path, file_name)
        if os.path.exists(file_path):
            solan = solan + 1
            file_name = str(Maso) + "_Lan_" + str(solan) + ".xlsx"      
        else:
            # Nếu file không tồn tại, tạo file mới
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Result_Test"
            wb.save(file_path)
            break
    C12 = "-"   # Impedance path 19v2 
    C13 = "-"   # Impedance path 7v2
    C14 = "-"   # Impedance path 5v2
    C15 = "-"   # Impedance path 5v0
    C16 = "-"   # Impedance path 3v3
    C17 = "-"   # Impedance path BAT+
    
    C18 = "-"   # Voltage path BAT+ 19v2 
    C19 = "-"   # Voltage path BAT+ 7v2
    C20 = "-"   # Voltage path BAT+ 5v2
    C21 = "-"   # Voltage path BAT+ 5v0
    C22 = "-"   # Voltage path BAT+ 3v3
    C23 = "-"   # Voltage path 12v0 BAT+ input
    
    C24 = "-"   # Voltage path CH+ 5v0
    C25 = "-"   # Voltage path CH+ 3v3

    C26 = "-"   # Load test program
    
    C27 = "-"   # RS485
    C28 = "-"   # Calib ADC
    C29 = "-"   # Test Mosfet
    C30 = "-"   # Test NTC
    C31 = "-"   # Test FAN
    
    C32 = "-"   # Load main program
    
    D28 = ""
    D29 = ""
    
    # Gửi lệnh reset mạch JIG-----------------------------------------------------
    _uart_var.send_data(cmd_reset)
    time.sleep(0.5)
    
    vErrorCode = ""
    
    # Gửi lệnh kiểm tra ID mạch cần test
    _uart_var.send_data(cmd_phase1_10)
       
    while True:  
        key = keyboard.is_pressed('q')
        if key:
            vErrorCode = "LỖI QUY TRÌNH TEST"
            print("Đã nhấn ESC. Kết thúc test mạch!")
            break   
            
        vKetthuc = False
        # Đọc liên tục từ cổng COM
        data_res = _uart_var.read_data()
        if data_res: 
            # print(data_res) 
            # print(len(data_res))
            CS_byte = 0 
            CS_byte = Cal_Checksum(data_res)                   
            if (data_res[0] == 0xEA) & (data_res[11] == CS_byte):
                # ID-----------------------------------------------------------------------
                if data_res[1] == CMD_CHECK_ID_CT:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> ID mạch cần test đúng !\n"
                        In_ketqua(Array_print)
                        print("> Đang test trở kháng.....")
                    else:
                        Array_print = "-> ERROR: Sai ID !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_ID_Power"
                        vKetthuc = True 
                # Impedance-----------------------------------------------------------------------
                elif data_res[1] == CMD_CHECK_IMPEDANCE:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test trở kháng thành công!\n"
                        In_ketqua(Array_print)
                        C12 = "PASS"   # Impedance path 19v2 
                        C13 = "PASS"   # Impedance path 7v2
                        C14 = "PASS"   # Impedance path 5v2
                        C15 = "PASS"   # Impedance path 5v0
                        C16 = "PASS"   # Impedance path 3v3
                        C17 = "PASS"   # Impedance path BAT+
                        print("> Đang test điện áp lộ BAT+.....")
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch POWER lộ 19v2 !!!\n"
                        In_loi(Array_print)
                        C12 = "ERROR"   # Impedance path 19v2
                        vErrorCode = "ERROR_TroKhang_19v2_Power"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch POWER lộ 7v2 !!!\n"
                        In_loi(Array_print)
                        C12 = "PASS"   # Impedance path 19v2 
                        C13 = "ERROR"   # Impedance path 7v2
                        vErrorCode = "ERROR_TroKhang_7v2_Power"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_3:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch POWER lộ 5v2 !!!\n"
                        In_loi(Array_print)
                        C12 = "PASS"   # Impedance path 19v2 
                        C13 = "PASS"   # Impedance path 7v2
                        C14 = "ERROR"   # Impedance path 5v2
                        vErrorCode = "ERROR_TroKhang_5v2_Power"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_4:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch POWER lộ 5v0 !!!\n"
                        In_loi(Array_print)
                        C12 = "PASS"   # Impedance path 19v2 
                        C13 = "PASS"   # Impedance path 7v2
                        C14 = "PASS"   # Impedance path 5v2
                        C15 = "ERROR"   # Impedance path 5v0
                        vErrorCode = "ERROR_TroKhang_5v0_Power"
                        vKetthuc = True   
                    elif data_res[2] == TEST_ERROR_5:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch POWER lộ 3v3 !!!\n"
                        In_loi(Array_print)
                        C12 = "PASS"   # Impedance path 19v2 
                        C13 = "PASS"   # Impedance path 7v2
                        C14 = "PASS"   # Impedance path 5v2
                        C15 = "PASS"   # Impedance path 5v0
                        C16 = "ERROR"   # Impedance path 3v3
                        vErrorCode = "ERROR_TroKhang_3v3_Power"
                        vKetthuc = True 
                    elif data_res[2] == TEST_ERROR_6:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch POWER lộ BAT+ (8v0) !!!\n"
                        In_loi(Array_print)
                        C12 = "PASS"   # Impedance path 19v2 
                        C13 = "PASS"   # Impedance path 7v2
                        C14 = "PASS"   # Impedance path 5v2
                        C15 = "PASS"   # Impedance path 5v0
                        C16 = "PASS"   # Impedance path 3v3
                        C17 = "ERROR"   # Impedance path BAT+
                        vErrorCode = "ERROR_TroKhang_BAT+_Power"
                        vKetthuc = True 
                # Voltage BAT+ -----------------------------------------------------------------------
                elif data_res[1] == CMD_CHECK_VOLTAGE_BAT:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test điện áp BAT+ thành công!\n"
                        In_ketqua(Array_print)
                        C18 = "PASS"   # Voltage path BAT+ 19v2 
                        C19 = "PASS"   # Voltage path BAT+ 7v2
                        C20 = "PASS"   # Voltage path BAT+ 5v2
                        C21 = "PASS"   # Voltage path BAT+ 5v0
                        C22 = "PASS"   # Voltage path BAT+ 3v3
                        C23 = "PASS"   # Voltage path 12v0 BAT+ input
                        print("> Đang test điện áp lộ CH+.....")
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch POWER lộ BAT+ 19v2 !!!\n"
                        In_loi(Array_print)
                        C18 = "ERROR"   # Voltage path BAT+ 19v2 
                        vErrorCode = "ERROR_DienApBAT_19v2_Power"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch POWER lộ BAT+ 7v2 !!!\n"
                        In_loi(Array_print)
                        C18 = "PASS"   # Voltage path BAT+ 19v2 
                        C19 = "ERROR"   # Voltage path BAT+ 7v2
                        vErrorCode = "ERROR_DienApBAT_7v2_Power"
                        vKetthuc = True  
                    elif data_res[2] == TEST_ERROR_3:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch POWER lộ BAT+ 5v2 !!!\n"
                        In_loi(Array_print)
                        C18 = "PASS"   # Voltage path BAT+ 19v2 
                        C19 = "PASS"   # Voltage path BAT+ 7v2
                        C20 = "ERROR"   # Voltage path BAT+ 5v2
                        vErrorCode = "ERROR_DienApBAT_5v2_Power"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_4:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch POWER lộ BAT+ 5v0 !!!\n"
                        In_loi(Array_print)
                        C18 = "PASS"   # Voltage path BAT+ 19v2 
                        C19 = "PASS"   # Voltage path BAT+ 7v2
                        C20 = "PASS"   # Voltage path BAT+ 5v2
                        C21 = "ERROR"   # Voltage path BAT+ 5v0
                        vErrorCode = "ERROR_DienApBAT_5v0_Power"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_5:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch POWER lộ BAT+ 3v3 !!!\n"
                        In_loi(Array_print)
                        C18 = "PASS"   # Voltage path BAT+ 19v2 
                        C19 = "PASS"   # Voltage path BAT+ 7v2
                        C20 = "PASS"   # Voltage path BAT+ 5v2
                        C21 = "PASS"   # Voltage path BAT+ 5v0
                        C22 = "ERROR"   # Voltage path BAT+ 3v3
                        vErrorCode = "ERROR_DienApBAT_3v3_Power"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_6:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch POWER lộ BAT+ 8v0 input !!!\n"
                        In_loi(Array_print)
                        C18 = "PASS"   # Voltage path BAT+ 19v2 
                        C19 = "PASS"   # Voltage path BAT+ 7v2
                        C20 = "PASS"   # Voltage path BAT+ 5v2
                        C21 = "PASS"   # Voltage path BAT+ 5v0
                        C22 = "PASS"   # Voltage path BAT+ 3v3
                        C23 = "ERROR"   # Voltage path 12v0 BAT+ input
                        vErrorCode = "ERROR_DienApBAT_BAT+_Power"
                        vKetthuc = True         
                # Voltage CH+ -----------------------------------------------------------------------
                elif data_res[1] == CMD_CHECK_VOLTAGE_CH:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test điện áp CH+ thành công!\n"
                        In_ketqua(Array_print)
                        C24 = "PASS"   # Voltage path CH+ 5v0
                        C25 = "PASS"   # Voltage path CH+ 3v3
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch POWER lộ CH+ 5v0 !!!\n"
                        In_loi(Array_print)
                        C24 = "ERROR"   # Voltage path CH+ 5v0
                        vErrorCode = "ERROR_DienApCH_5v0_Power"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch POWER lộ Ch+ 3v3 !!!\n"
                        In_loi(Array_print)
                        C24 = "PASS"   # Voltage path CH+ 5v0
                        C25 = "ERROR"   # Voltage path CH+ 3v3
                        vErrorCode = "ERROR_DienApCH_3v3_Power"
                        vKetthuc = True          
                # St_link-----------------------------------------------------------------------
                elif data_res[1] == CMD_CHECK_ST_LINK:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Đã kết nối St-Link, bắt đầu nạp code test !\n"
                        In_ketqua(Array_print)
                        time.sleep(0.5)
                        
                        # Nap code test
                        firm_path_center_test = firm_path_goc + "\power_board\power_test.hex"
                        cnt = 1
                        while cnt > 0:
                            kq_fw = flash_firmware(firm_path_center_test)
                            if kq_fw == 1:
                                break
                            cnt = cnt - 1
                            print("-> Nap lai code")
                        if kq_fw == 1:
                            Array_print = "-> Nạp firmware test thành công!\n"
                            In_ketqua(Array_print)
                            print("> Đang kiểm tra code test.....")
                            # Gửi lệnh kiểm tra nạp code test
                            time.sleep(0.1)
                            _uart_var.send_data(cmd_phase2_50)                            
                        else:
                            Array_print = "-> ERROR: Nạp firmware test thất bại !!!\n"
                            In_loi(Array_print)
                            C26 = "ERROR"   # Load test program
                            vErrorCode = "ERROR_NapCode_Power"
                            vKetthuc = True
                    else:
                        Array_print = "-> ERROR: Lỗi kết nối St-Link !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_STLink_Power"
                        vKetthuc = True                
                # Code test-----------------------------------------------------------------------
                elif data_res[1] == CMD_CHECK_CODE_TEST:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Kiểm tra nạp code test thành công!\n"
                        In_ketqua(Array_print)
                        C26 = "PASS"   # Load test program
                        print("> Đang test functions.....")
                    elif data_res[2] == TEST_ERROR:
                        Array_print = "-> ERROR: Kiểm tra nạp code test thất bại !!!\n"
                        In_loi(Array_print)
                        C26 = "ERROR"   # Load test program
                        print(Fore.MAGENTA+"             Bạn có muốn nạp lại không?      \n"+ Style.RESET_ALL)
                        print("          1. Có                    2. Không  \n")
                        kq = int(input("> Nhập:"))
                        if kq == 1:
                            # Nap lại code test
                            firm_path_center_test = firm_path_goc + "\power_board\power_test.hex"
                            kq_fw = flash_firmware(firm_path_center_test)
                            if kq_fw == 1:
                                Array_print = "-> Nạp firmware test thành công!\n"
                                In_ketqua(Array_print)
                                print("> Đang kiểm tra code test.....")
                                # Gửi lệnh kiểm tra lại nạp code test
                                time.sleep(0.1)
                                _uart_var.send_data(cmd_phase2_50)                                
                            else:
                                Array_print = "-> ERROR: Nạp firmware test thất bại !!!\n"
                                In_loi(Array_print)
                                C26 = "ERROR"   # Load test program
                                # Gửi lệnh reset mạch Jig
                                time.sleep(0.5)
                                _uart_var.send_data(cmd_reset)
                                vErrorCode = "ERROR_NapCode_Power"
                                vKetthuc = True
                        else:
                            # Gửi lệnh reset mạch Jig
                            time.sleep(0.5)
                            _uart_var.send_data(cmd_reset)
                            vErrorCode = "ERROR_NapCodeTest_Power"
                            vKetthuc = True              
                
                # Function-----------------------------------------------------------------------
                elif data_res[1] == CMD_CHECK_FUNCTIONS:
                    if data_res[2] == TEST_OK:
                        C27 = "PASS"   # RS485
                        C28 = "PASS"   # Calib ADC
                        C29 = "PASS"   # Test Mosfet
                        C30 = "PASS"   # Test NTC
                        C31 = "PASS"   # Test FAN

                        Array_print = "-> Test Funtion thành công, bắt đầu nạp code chính !\n"
                        In_ketqua(Array_print)
                        # Nạp code chính
                        firm_path_center_test = firm_path_goc + "\power_board\power_main.hex"
                        kq_fw = flash_firmware(firm_path_center_test)
                        if kq_fw == 1:
                            Array_print = "-> Nạp firmware chính thành công!\n"
                            In_ketqua(Array_print)
                            # Gửi lệnh kiểm tra nạp code chính
                            print("> Đang kiểm tra code chính.....") 
                            time.sleep(0.1)
                            _uart_var.send_data(cmd_phase3_70)  
                        else:
                            Array_print = "-> ERROR: Nạp firmware chính thất bại !!!\n"
                            In_loi(Array_print)
                            C32= "ERROR"   # Load main program
                            vErrorCode = "ERROR_NapCode_Power"
                            vKetthuc = True
                    # Rs485
                    elif data_res[2] == ERROR_RS485:
                        Array_print = "-> ERROR: Lỗi Mạch POWER RS485 !!!\n"
                        In_loi(Array_print)
                        C27 = "ERROR"   # RS485
                        vErrorCode = "ERROR_RS485_Power"
                        vKetthuc = True
                     
                    # Calib ADC   
                    elif data_res[2] == ERROR_OS_ADC_1_1:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_1_1 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_1_1_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_2_1:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_2_1 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_2_1_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_3_1:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_3_1 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_3_1_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_4_1:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_4_1 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_4_1_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_1_2:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_1_2 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_1_2_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_2_2:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_2_2 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_2_2_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_3_2:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_3_2 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_3_2_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_4_2:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_4_2 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_4_2_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_1_3:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_1_3 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_1_3_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_2_3:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_2_3 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_2_3_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_3_3:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_3_3 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_3_3_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_OS_ADC_4_3:
                        Array_print = "-> ERROR: Lỗi Mạch POWER Calib ADC offset_4_3 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "ERROR"   # Calib ADC
                        vErrorCode = "ERROR_ADC_OS_4_3_Power"
                        D28 = vErrorCode
                        vKetthuc = True
                    
                    # Test Mosfet
                    elif data_res[2] == ERROR_MOSFET_1:
                        Array_print = "-> ERROR: Lỗi Mạch POWER MOSFET 1 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "PASS"   # Calib ADC
                        C29 = "ERROR"   # Test Mosfet
                        vErrorCode = "ERROR_MOSFET1_Power"
                        D29 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_MOSFET_2:
                        Array_print = "-> ERROR: Lỗi Mạch POWER MOSFET 2 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "PASS"   # Calib ADC
                        C29 = "ERROR"   # Test Mosfet
                        vErrorCode = "ERROR_MOSFET2_Power"
                        D29 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_MOSFET_3:
                        Array_print = "-> ERROR: Lỗi Mạch POWER MOSFET 3 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "PASS"   # Calib ADC
                        C29 = "ERROR"   # Test Mosfet
                        vErrorCode = "ERROR_MOSFET3_Power"
                        D29 = vErrorCode
                        vKetthuc = True
                    elif data_res[2] == ERROR_MOSFET_4:
                        Array_print = "-> ERROR: Lỗi Mạch POWER MOSFET 4 !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "PASS"   # Calib ADC
                        C29 = "ERROR"   # Test Mosfet
                        vErrorCode = "ERROR_MOSFET4_Power"
                        D29 = vErrorCode
                        vKetthuc = True
                        
                    # Test NTC
                    elif data_res[2] == ERROR_NTC:
                        Array_print = "-> ERROR: Lỗi Mạch POWER NTC !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "PASS"   # Calib ADC
                        C29 = "PASS"   # Test Mosfet
                        C30 = "ERROR"   # Test NTC
                        vErrorCode = "ERROR_NTC_Power"
                        vKetthuc = True
                        
                    # Test FAN
                    elif data_res[2] == ERROR_FAN:
                        Array_print = "-> ERROR: Lỗi Mạch POWER FAN !!!\n"
                        In_loi(Array_print)
                        C27 = "PASS"   # RS485
                        C28 = "PASS"   # Calib ADC
                        C29 = "PASS"   # Test Mosfet
                        C30 = "PASS"   # Test NTC
                        C31 = "ERROR"   # Test FAN
                        vErrorCode = "ERROR_FAN_Power"
                        vKetthuc = True              
     
                # Code chính-----------------------------------------------------------------------
                elif data_res[1] == CMD_CHECK_CODE_MAIN:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Kiểm tra nạp code chính thành công, ***MẠCH ĐẠT YÊU CẦU!\n"
                        In_ketqua(Array_print)
                        C32 = "PASS"   # Load main program
                        vErrorCode = "TEST_POWER_OK"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR:
                        Array_print = "-> ERROR: Kiểm tra nạp code chính thất bại !!!\n"
                        In_loi(Array_print)
                        C32 = "ERROR"   # Load main program
                        print(Fore.MAGENTA +"             Bạn có muốn nạp lại không?      \n"+ Style.RESET_ALL)
                        print("          1. Có                    2. Không")
                        kq = int(input("> Nhập: "))
                        if kq == 1:
                            # Nạp lại code chính
                            firm_path_center_test = firm_path_goc + "\power_board\power_main.hex"
                            kq_fw = flash_firmware(firm_path_center_test)
                            if kq_fw == 1:
                                Array_print = "-> Nạp firmware chính thành công!\n"
                                In_ketqua(Array_print)
                                print("> Đang kiểm tra code chính.....")
                                # Gửi lệnh kiểm tra lại nạp code chính
                                time.sleep(0.1)
                                _uart_var.send_data(cmd_phase3_70)                                
                            else:
                                Array_print = "-> ERROR: Nạp firmware chính thất bại !!!\n"
                                In_loi(Array_print)
                                C32 = "ERROR"   # Load main program
                                vErrorCode = "ERROR_NapCode_Power"
                                vKetthuc = True
                        else:
                            # Gửi lệnh Reset mạch Jig
                            time.sleep(0.5)
                            _uart_var.send_data(cmd_reset)
                            vErrorCode = "ERROR_NapCodeChinh_Power"
                            vKetthuc = True
                else: #-----------------------------------------------------------------------
                    print(data_res)
                    print("Nothing\n")
            else:
                print("Size = " + len(data_res))
                print(data_res)
                print("-> Sai Cấu trúc bản tin Response !!!\n")
           
            
        if vKetthuc == True:
            break
    
        if keyboard.is_pressed('q'):
            vErrorCode = "LỖI QUY TRÌNH TEST"
            print("\n")
            print(Fore.RED + "> Đã nhấn Exit. Kết thúc test mạch!\n" + Style.RESET_ALL)
            break       
    
    # Lưu dữ liệu kết quả test-------------------------------------------------------
    wb = load_workbook(file_path)
    ws = wb.active
    
    ws["A1"] = "Board's name"
    
    if vErrorCode != "ERROR_ID_Power":
        ws["B1"] = "PWR - Power"
    else:
        ws["B1"] = "Sai loại mạch"
        ws["B1"].fill = red_fill
        
    ws["A2"] = "Design Company"
    ws["B2"] = "Viettel Post"
    ws["A3"] = "Factory's name"
    ws["B3"] = "Meiko"
    ws["A4"] = "Production Batch"
    ws["B4"] = LoSanXuat
    ws["A5"] = "Serial Number"
    ws["B5"] = Maso
    ws["A6"] = "Date"
    ws["A7"] = "Time"
        
    now = datetime.now()
    # Định dạng ngày và giờ
    ws["B6"] = now.strftime("%Y-%m-%d")  # Ngày hiện tại (YYYY-MM-DD)
    ws["B7"] = now.strftime("%H:%M:%S")  # Giờ hiện tại (HH:MM:SS)    
       
    ws["A8"] = "Number of times" 
    ws["B8"] =  solan
      
    ws["A9"] = "Result"   
    if vErrorCode == "TEST_POWER_OK":
       ws["B9"] =  "PASS"
       ws["B9"].fill =  green_fill
    else:
       ws["B9"] =  "ERROR" 
       ws["B9"].fill =  red_fill
    ws["B9"].font =  bold_font
    
    ws["A11"] =  "No" 
    ws["B11"] =  "Test Function" 
    ws["C11"] =  "Result"
    ws["D11"] =  "Note"
    
    ws["A12"] =  '1'
    ws["A13"] =  '2'
    ws["A14"] =  '3'
    ws["A15"] =  '4'
    ws["A16"] =  '5'
    ws["A17"] =  '6'
    ws["A18"] =  '7'
    ws["A19"] =  '8'
    ws["A20"] =  '9'
    ws["A21"] =  '10'
    ws["A22"] =  '11'
    ws["A23"] =  '12'
    ws["A24"] =  '13'
    ws["A25"] =  '14'
    ws["A26"] =  '15'
    ws["A27"] =  '16'
    ws["A28"] =  '17'
    ws["A29"] =  '18'
    ws["A30"] =  '19'
    ws["A31"] =  '20'
    ws["A32"] =  '21'
    
    ws["B12"] =  "Impedance path 19v2"
    ws["B13"] =  "Impedance path 7v2"
    ws["B14"] =  "Impedance path 5v2"
    ws["B15"] =  "Impedance path 5v0"
    ws["B16"] =  "Impedance path 3v3"
    ws["B17"] =  "Impedance path BAT+"
    ws["B18"] =  "Voltage path BAT+ 19v2"
    ws["B19"] =  "Voltage path BAT+ 7v2"
    ws["B20"] =  "Voltage path BAT+ 5v2"
    ws["B21"] =  "Voltage path BAT+ 5v0"
    ws["B22"] =  "Voltage path BAT+ 3v3"
    ws["B23"] =  "Voltage path BAT+ BAT+"
    ws["B24"] =  "Voltage path CH+ 5v0"
    ws["B25"] =  "Voltage path CH+ 3v3"
    ws["B26"] =  "Load test program"
    ws["B27"] =  "RS485"
    ws["B28"] =  "Calib ADC"
    ws["B29"] =  "Test Mosfet"
    ws["B30"] =  "Test NTC"
    ws["B31"] =  "Test FAN"
    ws["B32"] =  "Load main program"
    
    ws["C12"] =  C12
    ws["C13"] =  C13
    ws["C14"] =  C14
    ws["C15"] =  C15
    ws["C16"] =  C16
    ws["C17"] =  C17
    ws["C18"] =  C18
    ws["C19"] =  C19
    ws["C20"] =  C20
    ws["C21"] =  C21
    ws["C22"] =  C22
    ws["C23"] =  C23
    ws["C24"] =  C24
    ws["C25"] =  C25
    ws["C26"] =  C26
    ws["C27"] =  C27
    ws["C28"] =  C28
    ws["C29"] =  C29
    ws["C30"] =  C30
    ws["C31"] =  C31
    ws["C32"] =  C32
    
    ws["D28"] =  D28
    ws["D29"] =  D29
    
    # Lưu file Excel
    wb.save(file_path)
    
    # Gửi lệnh reset Jig--------------------------------------------------------------------    
    time.sleep(0.1)
    _uart_var.send_data(cmd_reset)
    return vErrorCode