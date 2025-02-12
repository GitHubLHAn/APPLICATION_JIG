import time
import keyboard

from uart import *

from colorama import Fore, Style

from loadcode import *

import os
from openpyxl import load_workbook
from openpyxl import Workbook

from datetime import datetime

from openpyxl.styles import PatternFill
from openpyxl.styles import Font

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
bold_font = Font(bold=True)  # Định dạng chữ đậm


TEST_OK         =   0x59
TEST_ERROR  	=	0x4E
TEST_ERROR_1	=	0x5E
TEST_ERROR_2	=	0x6E
TEST_ERROR_3	=	0x7E
TEST_ERROR_4	=	0x8E

CMD_CHECK_ID_CT		    =	0x10
CMD_CHECK_IMPEDANCE	    =   0x20
CMD_CHECK_VOLTAGE	    =	0x30
CMD_CHECK_ST_LINK	    =	0x40
CMD_CHECK_CODE_TEST	    =   0x50
CMD_CHECK_FUNCTIONS	    =   0x60
CMD_CHECK_CODE_MAIN     =	0x70

Array_print = "Mang in ra"

def In_loi(data):
    print(Fore.RED + data + Style.RESET_ALL)
    
def In_ketqua(data):
    print(Fore.GREEN + data + Style.RESET_ALL)
    
def Cal_Checksum(data):
    CS_b = 0
    for i in range(0,11):
        CS_b += data[i]
        CS_b = CS_b % 256
    # CS_b = data[0] + data[1] + data[2] + data[3] + data[4] + data[5] + data[6] + data[7] + data[8] + data[9] + data[10] + data[11]
    # CS_b = CS_b % 256
    return CS_b

def Sensor_process(_uart_var, firm_path_goc, log_path_goc, infor_mach):
    # Lưu log dữ liệu test của mạch ------------------------------------------ 
    folder_path = log_path_goc+ "\Sensor"

    part_infor = infor_mach.split('_')
    Header = part_infor[0]
    MachCanTest = part_infor[1]
    LoSanXuat = part_infor[2]
    Maso = part_infor[3]
    file_name = Maso + ".xlsx"
    
    solan = 1
    while True:
        file_path = os.path.join(folder_path, file_name)
        if os.path.exists(file_path):
            solan = solan + 1
            file_name = str(Maso) + "_Lan_" + str(solan) + ".xlsx"      
        else:
            # Nếu file không tồn tại, tạo file mới
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Result_Test"
            wb.save(file_path)
            break
    C12 = "-"   # Impedance path 8v0 
    C13 = "-"   # Impedance path 5vBus
    C14 = "-"   # Impedance path 5v0
    C15 = "-"   # Impedance path 3v3
    C16 = "-"   # Voltage path 5v0
    C17 = "-"   # Voltage path 3v3
    C18 = "-"   # Load test program
    C19 = "-"   # RS485
    C20 = "-"   # Hall Sensors
    C21 = "-"   # RFID
    C22 = "-"   # Load main program
    
    # Gửi lệnh reset mạch JIG-----------------------------------------------------
    buffer_cmd = [0xAB, 0xCD, 0x00, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x23]
    _uart_var.send_data(buffer_cmd)
    time.sleep(0.5)
    
    vErrorCode = ""
    
    # Gửi lệnh kiểm tra ID mạch cần test
    buffer_cmd = [0xAB, 0xCD, 0x10, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x33]
    _uart_var.send_data(buffer_cmd)
    
    while True:      
        vKetthuc = False
        # Đọc liên tục từ cổng COM
        data_res = _uart_var.read_data()
        if data_res: 
            # print(data_res) 
            # print(len(data_res))
            CS_byte = 0 
            CS_byte = Cal_Checksum(data_res)                   
            if (data_res[0] == 0xEA) & (data_res[11] == CS_byte):
                # ID
                if data_res[1] == CMD_CHECK_ID_CT:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> ID mạch cần test đúng !\n"
                        In_ketqua(Array_print)
                        print("> Đang test trở kháng.....")
                    else:
                        Array_print = "-> ERROR: Sai ID !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_ID_Sensor"
                        vKetthuc = True
                # Impedance
                elif data_res[1] == CMD_CHECK_IMPEDANCE:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test trở kháng thành công!\n"
                        In_ketqua(Array_print)
                        C12 = "PASS"   # Impedance path 8v0 
                        C13 = "PASS"   # Impedance path 5vBus
                        C14 = "PASS"   # Impedance path 5v0
                        C15 = "PASS"   # Impedance path 3v3
                        print("> Đang test điện áp.....")
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch SENSOR lộ 8v0 !!!\n"
                        In_loi(Array_print)
                        C12 = "ERROR"   # Impedance path 8v0 
                        vErrorCode = "ERROR_TroKhang_8v0_Sensor"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch SENSOR lộ 5vBus !!!\n"
                        In_loi(Array_print)
                        C12 = "PASS"   # Impedance path 8v0 
                        C13 = "ERROR"   # Impedance path 5vBus
                        vErrorCode = "ERROR_TroKhang_5vBus_Sensor"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_3:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch SENSOR lộ 5v0 !!!\n"
                        In_loi(Array_print)
                        C12 = "PASS"   # Impedance path 8v0 
                        C13 = "PASS"   # Impedance path 5vBus
                        C14 = "ERROR"   # Impedance path 5v0
                        vErrorCode = "ERROR_TroKhang_5v0_Sensor"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_4:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch SENSOR lộ 3v3 !!!\n"
                        In_loi(Array_print)
                        C12 = "PASS"   # Impedance path 8v0 
                        C13 = "PASS"   # Impedance path 5vBus
                        C14 = "PASS"   # Impedance path 5v0
                        C15 = "ERROR"   # Impedance path 3v3
                        vErrorCode = "ERROR_TroKhang_3v3_Sensor"
                        vKetthuc = True
                 
                # Voltage
                elif data_res[1] == CMD_CHECK_VOLTAGE:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test điện áp thành công!\n"
                        In_ketqua(Array_print)
                        C16 = "PASS"   # Voltage path 5v0
                        C17 = "PASS"   # Voltage path 3v3
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch SENSOR lộ 5v0 !!!\n"
                        In_loi(Array_print)
                        C16 = "ERROR"   # Voltage path 5v0
                        vErrorCode = "ERROR_DienAp_5v0_Sensor"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch SENSOR lộ 3v3 !!!\n"
                        In_loi(Array_print)
                        C16 = "PASS"   # Voltage path 5v0
                        C17 = "ERROR"   # Voltage path 3v3
                        vErrorCode = "ERROR_DienAp_3v3_Sensor"
                        vKetthuc = True
                # St_link
                elif data_res[1] == CMD_CHECK_ST_LINK:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Đã kết nối St-Link, bắt đầu nạp code test !\n"
                        In_ketqua(Array_print)
                        time.sleep(0.5)
                        # Nap code test
                        firm_path_sensor_test = firm_path_goc + "\sensor_board\sensor_test.hex"
                        kq_fw = flash_firmware(firm_path_sensor_test)
                        if kq_fw == 1:
                            Array_print = "-> Nạp firmware test thành công!\n"
                            In_ketqua(Array_print)
                            # Gửi lệnh kiểm tra nạp code test
                            time.sleep(1)
                            buffer_cmd = [0xAB, 0xCD, 0x50, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x73]
                            _uart_var.send_data(buffer_cmd)
                            print("> Đang kiểm tra code test.....")
                        else:
                            Array_print = "-> ERROR: Nạp firmware test thất bại !!!\n"
                            In_loi(Array_print)
                            C18 = "ERROR"   # Load test program
                            vErrorCode = "ERROR_NapCode_Sensor"
                            vKetthuc = True
                    else:
                        Array_print = "-> ERROR: Lỗi kết nối St-Link !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_STLink_Sensor"
                        vKetthuc = True
                # Code test
                elif data_res[1] == CMD_CHECK_CODE_TEST:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Kiểm tra nạp code test thành công!\n"
                        In_ketqua(Array_print)
                        C18 = "PASS"   # Load test program
                        print("> Đang test functions.....")
                    elif data_res[2] == TEST_ERROR:
                        Array_print = "-> ERROR: Kiểm tra nạp code test thất bại !!!\n"
                        In_loi(Array_print)
                        C18 = "ERROR"   # Load test program
                        print("             Bạn có muốn nạp lại không?      \n")
                        print("          1. Có                    2. Không  \n")
                        kq = int(input("> Nhập:"))
                        if kq == 1:
                            # Nap lại code test
                            firm_path_sensor_test = firm_path_goc + "\sensor_board\sensor_test.hex"
                            kq_fw = flash_firmware(firm_path_sensor_test)
                            if kq_fw == 1:
                                Array_print = "-> Nạp firmware test thành công!\n"
                                In_ketqua(Array_print)
                                # Gửi lệnh kiểm tra lại nạp code test
                                time.sleep(1)
                                buffer_cmd = [0xAB, 0xCD, 0x50, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x73]
                                _uart_var.send_data(buffer_cmd)
                                print("> Đang kiểm tra code test.....")
                            else:
                                Array_print = "-> ERROR: Nạp firmware test thất bại !!!\n"
                                In_loi(Array_print)
                                C18 = "ERROR"   # Load test program
                                # Gửi lệnh reset mạch Jig
                                time.sleep(0.5)
                                buffer_cmd = [0xAB, 0xCD, 0x00, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x23]
                                _uart_var.send_data(buffer_cmd)
                                vErrorCode = "ERROR_NapCode_Sensor"
                                vKetthuc = True
                        else:
                            # Gửi lệnh reset mạch Jig
                            time.sleep(0.5)
                            buffer_cmd = [0xAB, 0xCD, 0x00, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x23]
                            _uart_var.send_data(buffer_cmd)
                            vErrorCode = "ERROR_NapCodeTest_Sensor"
                            vKetthuc = True
                # Function
                elif data_res[1] == CMD_CHECK_FUNCTIONS:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test Funtion thành công, bắt đầu nạp code chính !\n"
                        In_ketqua(Array_print)
                        C19 = "PASS"   # RS485
                        C20 = "PASS"   # Hall Sensors
                        C21 = "PASS"   # RFID
                        # Nạp code chính
                        firm_path_sensor_test = firm_path_goc + "\sensor_board\sensor_main.hex"
                        kq_fw = flash_firmware(firm_path_sensor_test)
                        if kq_fw == 1:
                            Array_print = "-> Nạp firmware chính thành công!\n"
                            In_ketqua(Array_print)
                            # Gửi lệnh kiểm tra nạp code chính
                            time.sleep(1)
                            buffer_cmd = [0xAB, 0xCD, 0x70, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x93]
                            _uart_var.send_data(buffer_cmd)  
                            print("> Đang kiểm tra code chính.....") 
                        else:
                            Array_print = "-> ERROR: Nạp firmware chính thất bại !!!\n"
                            In_loi(Array_print)
                            C22 = "ERROR"   # Load main program
                            vErrorCode = "ERROR_NapCode_Sensor"
                            vKetthuc = True
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi Mạch SENSOR RS485 !!!\n"
                        In_loi(Array_print)
                        C19 = "ERROR"   # RS485
                        vErrorCode = "ERROR_RS485_Sensor"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi Mạch SENSOR Hall !!!\n"
                        In_loi(Array_print)
                        C19 = "PASS"   # RS485
                        C20 = "ERROR"   # Hall Sensors
                        vErrorCode = "ERROR_Hall_Sensor"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR_3:
                        Array_print = "-> ERROR: Lỗi Mạch SENSOR RFID !!!\n"
                        In_loi(Array_print)
                        C19 = "PASS"   # RS485
                        C20 = "PASS"   # Hall Sensors
                        C21 = "ERROR"   # RFID
                        vErrorCode = "ERROR_RFID_Sensor"
                        vKetthuc = True
                # Code chính
                elif data_res[1] == CMD_CHECK_CODE_MAIN:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Kiểm tra nạp code chính thành công, ***MẠCH ĐẠT YÊU CẦU!\n"
                        In_ketqua(Array_print)
                        C22 = "PASS"   # Load main program
                        vErrorCode = "TEST_SENSOR_OK"
                        vKetthuc = True
                    elif data_res[2] == TEST_ERROR:
                        Array_print = "-> ERROR: Kiểm tra nạp code chính thất bại !!!\n"
                        In_loi(Array_print)
                        C22 = "ERROR"   # Load main program
                        print("             Bạn có muốn nạp lại không?      \n")
                        print("          1. Có                    2. Không")
                        kq = int(input("> Nhập: "))
                        if kq == 1:
                            # Nạp lại code chính
                            firm_path_sensor_test = firm_path_goc + "\sensor_board\sensor_main.hex"
                            kq_fw = flash_firmware(firm_path_sensor_test)
                            if kq_fw == 1:
                                Array_print = "-> Nạp firmware chính thành công!\n"
                                In_ketqua(Array_print)
                                # Gửi lệnh kiểm tra lại nạp code chính
                                time.sleep(1)
                                buffer_cmd = [0xAB, 0xCD, 0x70, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x93]
                                _uart_var.send_data(buffer_cmd)
                                print("> Đang kiểm tra code chính.....")
                            else:
                                Array_print = "-> ERROR: Nạp firmware chính thất bại !!!\n"
                                In_loi(Array_print)
                                C22 = "ERROR"   # Load main program
                                vErrorCode = "ERROR_NapCode_Sensor"
                                vKetthuc = True
                        else:
                            # Gửi lệnh Reset mạch Jig
                            time.sleep(0.5)
                            buffer_cmd = [0xAB, 0xCD, 0x00, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x23]
                            _uart_var.send_data(buffer_cmd)
                            vErrorCode = "ERROR_NapCodeChinh_Sensor"
                            vKetthuc = True
                else:
                    print(data_res)
                    print("Nothing\n")
            else:
                print(len(data_res))
                print("-> Sai Cấu trúc bản tin Response !!!\n")
           
            
        if vKetthuc == True:
            break
        
        if keyboard.is_pressed('q'):
            vErrorCode = "LỖI QUY TRÌNH TEST"
            print("\n")
            print(Fore.RED + "> Đã nhấn Exit. Kết thúc test mạch!\n" + Style.RESET_ALL)
            break      
        
    # Lưu dữ liệu kết quả test-------------------------------------------------------
    wb = load_workbook(file_path)
    ws = wb.active
    
    ws["A1"] = "Board's name"
    if vErrorCode != "ERROR_ID_Sensor":
        ws["B1"] = "SS - Sensor"
    else:
        ws["B1"] = "Sai loại mạch"
        ws["B1"].fill = red_fill
    ws["A2"] = "Design Company"
    ws["B2"] = "Viettel Post"
    ws["A3"] = "Factory's name"
    ws["B3"] = "Meiko"
    ws["A4"] = "Production Batch"
    ws["B4"] = LoSanXuat
    ws["A5"] = "Serial Number"
    ws["B5"] = Maso
    ws["A6"] = "Date"
    ws["A7"] = "Time"
        
    now = datetime.now()
    # Định dạng ngày và giờ
    ws["B6"] = now.strftime("%Y-%m-%d")  # Ngày hiện tại (YYYY-MM-DD)
    ws["B7"] = now.strftime("%H:%M:%S")  # Giờ hiện tại (HH:MM:SS)    
       
    ws["A8"] = "Number of times" 
    ws["B8"] =  solan
      
    ws["A9"] = "Result"   
    if vErrorCode == "TEST_SENSOR_OK":
       ws["B9"] =  "PASS"
       ws["B9"].fill =  green_fill
    else:
       ws["B9"] =  "ERROR" 
       ws["B9"].fill =  red_fill
    ws["B9"].font =  bold_font
    
    ws["A11"] =  "No" 
    ws["B11"] =  "Test Function" 
    ws["C11"] =  "Result"
    
    ws["A12"] =  '1'
    ws["A13"] =  '2'
    ws["A14"] =  '3'
    ws["A15"] =  '4'
    ws["A16"] =  '5'
    ws["A17"] =  '6'
    ws["A18"] =  '7'
    ws["A19"] =  '8'
    ws["A20"] =  '9'
    ws["A21"] =  '10'
    ws["A22"] =  '11'
    
    ws["B12"] =  "Impedance path 8v0"
    ws["B13"] =  "Impedance path 5vBus"
    ws["B14"] =  "Impedance path 5v0"
    ws["B15"] =  "Impedance path 3v3"
    ws["B16"] =  "Voltage path 5v0"
    ws["B17"] =  "Voltage path 3v3"
    ws["B18"] =  "Load test program"
    ws["B19"] =  "RS485"
    ws["B20"] =  "Hall Sensors"
    ws["B21"] =  "RRID Sensor"
    ws["B22"] =  "Load main program"
    
    ws["C12"] =  C12
    ws["C13"] =  C13
    ws["C14"] =  C14
    ws["C15"] =  C15
    ws["C16"] =  C16
    ws["C17"] =  C17
    ws["C18"] =  C18
    ws["C19"] =  C19
    ws["C20"] =  C20
    ws["C21"] =  C21
    ws["C22"] =  C22
    
    # Lưu file Excel
    wb.save(file_path)
    
    # Gửi lệnh reset Jig--------------------------------------------------------------------    
    time.sleep(1)
    buffer_cmd = [0xAB, 0xCD, 0x00, 0x05, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x23]
    _uart_var.send_data(buffer_cmd)
    return vErrorCode