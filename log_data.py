from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from openpyxl.styles import Font

from colorama import Fore, Style


def log_test_result_func(log_path_goc, infor, result, MachCanTest):
    if MachCanTest == 'BS':
        file_name  = "\Boost\Boost_result.xlsx"    
    elif MachCanTest == 'CT':   
        file_name  = "\Center\Center_result.xlsx"     
    elif MachCanTest == 'LED': 
        file_name  = "\Led\Led_result.xlsx"         
    elif MachCanTest == 'APR': 
        file_name  = "\APR\APR_result.xlsx"        
    elif MachCanTest == 'SS': 
        file_name  = "\Sensor\Sensor_result.xlsx"
    elif MachCanTest == 'PWR': 
        file_name  = "\Power\Power_result.xlsx"              
    elif MachCanTest == 'DC': 
        file_name  = "\DocCharge\Charge_result.xlsx"   
    else:
        print("> Sai loại mạch xxx")
        return -1
    
    print("> Đang lưu kết quả test...")
   
    parts_infor = infor.split('_')
    parts_result = result.split('_')

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")

    bold_font = Font(bold=True)  # Định dạng chữ đậm

    # Mở file Excel
    wb = load_workbook(log_path_goc + file_name)
    ws = wb.active

    # Bắt đầu ghi số thứ tự từ dòng thứ 2
    cnt_row = 2

    while True:
        cell = ws[f"A{cnt_row}"]  # Lấy ô trong cột A, hàng hiện tại
        # print(cell)
        
        if cell.value is None:  # Nếu ô trống
            if cnt_row == 2:
                # print("Bat dau ghi du lieu dau tien")
                cell.value = 1
                break
            else:
                # print("O con trong, them dui lieu")
                cell.value = pre_value + 1
                break
        else:
            cnt_row = cnt_row + 1
            pre_value = cell.value
        
    now = datetime.now()
    # Định dạng ngày và giờ
    current_date = now.strftime("%Y-%m-%d")  # Ngày hiện tại (YYYY-MM-DD)
    current_time = now.strftime("%H:%M:%S")  # Giờ hiện tại (HH:MM:SS)
    
    ws[f"B{cnt_row}"] = current_date
    ws[f"C{cnt_row}"] = current_time

    ws[f"D{cnt_row}"] = str(parts_infor[0])     # Đơn vị thiết kế
    ws[f"E{cnt_row}"] = str(parts_infor[1])     # Loại mạch
    ws[f"F{cnt_row}"] = str(parts_infor[2])     # Lô sản xuất
    ws[f"G{cnt_row}"] = str(parts_infor[3])     # Số serial 

    ws[f"H{cnt_row}"] = str(result)
    ws[f"H{cnt_row}"].font = bold_font

    if parts_result[0] == "ERROR":
        ws[f"H{cnt_row}"].fill = red_fill
    else:
        if result == "LỖI QUY TRÌNH TEST":
            ws[f"H{cnt_row}"].fill = purple_fill
        else:
            ws[f"H{cnt_row}"].fill = green_fill
        
    solantest = 1
    for i in range(2, cnt_row):
        seri_number_cell = ws[f"G{i}"]
        seri_number = str(seri_number_cell.value) 
        if seri_number == parts_infor[3]:
            solantest = solantest + 1


    if solantest > 1:
        ghichu = f"Test lần {solantest}"
        ws[f"I{cnt_row}"] = ghichu
        ws[f"I{cnt_row}"].fill = yellow_fill

    # Lưu file Excel
    wb.save(log_path_goc + file_name)
    print(Fore.GREEN + "-> Đã lưu kết quả test\n"+ Style.RESET_ALL)
    
    
    
    
# def log_test_result_boost(log_path_goc, infor, result):
    
    
# def log_test_result_center(log_path_goc, infor, result):






# def log_test_result_led(log_path_goc, infor, result):





# def log_test_result_APR(log_path_goc, infor, result):





# def log_test_result_sensor(log_path_goc, infor, result):





# def log_test_result_power(log_path_goc, infor, result):



# def log_test_result_doccharge(log_path_goc, infor, result):
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    