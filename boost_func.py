import time
import keyboard

from uart import *

from colorama import Fore, Style

import os
from openpyxl import load_workbook
from openpyxl import Workbook

from datetime import datetime

from openpyxl.styles import PatternFill
from openpyxl.styles import Font

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
bold_font = Font(bold=True)  # Định dạng chữ đậm


TEST_OK         =   0x59
TEST_ERROR  	=	0x4E
TEST_ERROR_1	=	0x5E
TEST_ERROR_2	=	0x6E
TEST_ERROR_3	=	0x7E
TEST_ERROR_4	=	0x8E

CMD_CHECK_ID_CT		    =	0x10
CMD_CHECK_IMPEDANCE	    =   0x20
CMD_CHECK_VOLTAGE	    =	0x30

CMD_buff_reset = [0xAE, 0x00, 0x01, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xFF]
CMD_buff_test = [0xAE, 0x10, 0x01, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0xAA, 0x0F]

Array_print = "Mang in ra"

def In_loi(data):
    print(Fore.RED + data + Style.RESET_ALL)
    
def In_ketqua(data):
    print(Fore.GREEN + data + Style.RESET_ALL)
    
def Cal_Checksum(data):
    # CS_b = 0
    # for i in range(0,11):
    #     CS_b += data[i]
    #     CS_b = CS_b % 256
    # return CS_b
    return sum(data[:11]) % 256

def Boost_process(_uart_var, log_path_goc, infor_mach):
     # Lưu log dữ liệu test của mạch ------------------------------------------ 
    folder_path = log_path_goc+ "\Boost"

    part_infor = infor_mach.split('_')
    Header = part_infor[0]
    MachCanTest = part_infor[1]
    LoSanXuat = part_infor[2]
    Maso = part_infor[3]
    file_name = Maso + ".xlsx"
    
    solan = 1
    while True:
        file_path = os.path.join(folder_path, file_name)
        if os.path.exists(file_path):
            solan = solan + 1
            file_name = str(Maso) + "_Lan_" + str(solan) + ".xlsx"

                
        else:
            # Nếu file không tồn tại, tạo file mới
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Result_Test"
            wb.save(file_path)
            # print(f"->File '{file_name}' đã được tạo.")
            break
    C12 = "-"
    C13 = "-"
    C14 = "-"
    C15 = "-"
    C16 = "-"
    C17 = "-"
    
    # Gửi lệnh reset mạch JIG-------------------------------------------------
    _uart_var.send_data(CMD_buff_reset)
    time.sleep(0.5)
    
    vErrorCode = ""
    
    # Gửi lệnh kiểm tra ID mạch cần test
    _uart_var.send_data(CMD_buff_test)
       
    while True:      
        vKetthuc = False
        # Đọc liên tục từ cổng COM
        data_res = _uart_var.read_data()
        if data_res: 
            # print(data_res) 
            # print(len(data_res))
            CS_byte = 0 
            CS_byte = Cal_Checksum(data_res)                   
            if (data_res[0] == 0xEA) & (data_res[11] == CS_byte):
                # ID
                if data_res[1] == CMD_CHECK_ID_CT:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> ID mạch cần test đúng !\n"
                        In_ketqua(Array_print)
                        print("> Đang test trở kháng.....")
                    else:
                        Array_print = "-> ERROR: Sai ID !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_ID_Boost"
                        vKetthuc = True 
                # Impedance
                elif data_res[1] == CMD_CHECK_IMPEDANCE:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test trở kháng thành công!\n"
                        In_ketqua(Array_print)
                        print("> Đang test điện áp.....")
                        C12 = "PASS"
                        C13 = "PASS"
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch BOOST lộ input !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_TroKhang_Input_Boost"
                        vKetthuc = True
                        C12 = "ERROR"
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi test trở kháng mạch BOOST lộ output !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_TroKhang_Output_Boost"
                        vKetthuc = True
                        C12 = "PASS"
                        C13 = "ERROR"
                # Voltage
                elif data_res[1] == CMD_CHECK_VOLTAGE:
                    if data_res[2] == TEST_OK:
                        Array_print = "-> Test điện áp thành công!\n"
                        In_ketqua(Array_print)
                        vErrorCode = "TEST_BOOST_OK"
                        vKetthuc = True  
                        C14 = "PASS"
                        C15 = "PASS"
                        C16 = "PASS"
                        C17 = "PASS"
                    elif data_res[2] == TEST_ERROR_1:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch BOOST lộ 12v0 đầu vào !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_DienAp_12v0_Boost"
                        vKetthuc = True
                        C14 = "PASS" 
                    elif data_res[2] == TEST_ERROR_2:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch BOOST lộ 36v0 đầu ra không tải !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_DienAp_36v0_Boost"
                        vKetthuc = True
                        C14 = "PASS"
                        C15 = "ERROR"         
                    elif data_res[2] == TEST_ERROR_3:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch BOOST đầu ra tải nhỏ !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_DienAp_TaiNho_Boost"
                        vKetthuc = True 
                        C14 = "PASS"
                        C15 = "PASS"
                        C16 = "ERROR"  
                    elif data_res[2] == TEST_ERROR_4:
                        Array_print = "-> ERROR: Lỗi test điện áp mạch BOOST đầu ra tải lớn !!!\n"
                        In_loi(Array_print)
                        vErrorCode = "ERROR_DienAp_TaiLon_Boost"
                        vKetthuc = True
                        C14 = "PASS"
                        C15 = "PASS"
                        C16 = "PASS"
                        C17 = "ERROR"      
                else:
                    print(data_res)
                    print("Nothing\n")
            else:
                print("Size = " + len(data_res))
                print(data_res)
                print("-> Sai Cấu trúc bản tin Response !!!\n")
           
            
        if vKetthuc == True:
            break
        
        if keyboard.is_pressed('q'):
            vErrorCode = "LỖI QUY TRÌNH TEST"
            print("\n")
            print(Fore.RED + "> Đã nhấn Exit. Kết thúc test mạch!\n" + Style.RESET_ALL)
            break     
        
    # vErrorCode = "TEST_BOOST_OK"
    
    # Lưu dữ liệu kết quả test
    wb = load_workbook(file_path)
    ws = wb.active
    
    ws["A1"] = "Board's name"
    ws["B1"] = "Boost"
    ws["A2"] = "Design Company"
    ws["B2"] = "Viettel Post"
    ws["A3"] = "Factory's name"
    ws["B3"] = "Meiko"
    ws["A4"] = "Production Batch"
    ws["B4"] = LoSanXuat
    ws["A5"] = "Serial Number"
    ws["B5"] = Maso
    ws["A6"] = "Date"
    ws["A7"] = "Time"
        
    now = datetime.now()
    # Định dạng ngày và giờ
    ws["B6"] = now.strftime("%Y-%m-%d")  # Ngày hiện tại (YYYY-MM-DD)
    ws["B7"] = now.strftime("%H:%M:%S")  # Giờ hiện tại (HH:MM:SS)    
       
    ws["A8"] = "Number of times" 
    ws["B8"] =  solan
      
    ws["A9"] = "Result"   
    if vErrorCode == "TEST_BOOST_OK":
       ws["B9"] =  "PASS"
       ws["B9"].fill =  green_fill
    else:
       ws["B9"] =  "ERROR" 
       ws["B9"].fill =  red_fill
    ws["B9"].font =  bold_font
    
    ws["A11"] =  "No" 
    ws["B11"] =  "Test Function" 
    ws["C11"] =  "Result"
    
    ws["A12"] =  '1'
    ws["A13"] =  '2'
    ws["A14"] =  '3'
    ws["A15"] =  '4'
    ws["A16"] =  '5'
    ws["A17"] =  '6'
    
    ws["B12"] =  "Impedance Input path"
    ws["B13"] =  "Impedance Output path"
    ws["B14"] =  "Voltage 12v0 Input"
    ws["B15"] =  "Voltage 36v0 Output no load"
    ws["B16"] =  "Voltage 36v0 Output low load"
    ws["B17"] =  "Voltage 36v0 Output high load"
    
    ws["C12"] =  C12
    ws["C13"] =  C13
    ws["C14"] =  C14
    ws["C15"] =  C15
    ws["C16"] =  C16
    ws["C17"] =  C17
    
    # Lưu file Excel
    wb.save(file_path)
        
    time.sleep(0.1)
    _uart_var.send_data(CMD_buff_reset)
    return vErrorCode